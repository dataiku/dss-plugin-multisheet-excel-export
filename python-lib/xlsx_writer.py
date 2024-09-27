#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
Utility functions for conversion to Excel.
Conversion is based on Pandas feature conversion to xlsx.
"""

import logging
import math
import os
import tempfile
import zipfile

from typing import Tuple, List, Dict
from copy import copy

from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Fill, Font
from openpyxl.styles.borders import Border
from openpyxl.styles.colors import WHITE
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from zipfile import ZIP_DEFLATED

DATAIKU_TEAL = "FF2AB1AC"
LETTER_WIDTH = 1.20 # Approximative letter width to scale column width
MAX_LENGTH_TO_SHOW = 45 # Limit copied from DSS native excel exporter
EXCEL_MAX_LEN_SHEET_NAME = 31

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='Multi-Sheet Excel Exporter | %(levelname)s - %(message)s')

def get_column_width(column: Tuple):
    """
    Find optimum column width based on content and header length
    Based on the computations of DSS native excel output formatter
    """

    header = column[0]
    length_header = len(str(header.value))

    sum_length_cells = 0
    max_length_cells = 0
    for cell in column:
        length_cell = len(str(cell.value))
        max_length_cells = max(max_length_cells, length_cell)
        sum_length_cells += length_cell

    # Computations from ExcelOutputFormatter.java ExcelOutputFormatter.footer
    average_length_cell = math.ceil(sum_length_cells / (len(column) + 1))
    max_length_cells = min(max_length_cells, MAX_LENGTH_TO_SHOW)
    
    if max_length_cells > 2 * average_length_cell: # if max length much bigger than average
        length_to_show = int((max_length_cells + average_length_cell) / 2)
    else:
        length_to_show = max_length_cells

    length_to_show = max(length_to_show, length_header) 

    return length_to_show * LETTER_WIDTH

def auto_size_column_width(worksheet: Worksheet):
    """
    Resize columns based on the length of the header text
    """
    if worksheet.min_column < 1:
        logger.warning(f"No header row for worksheet '{worksheet}'. Column auto-size skipped.")
        return

    dimension_holder = DimensionHolder(worksheet=worksheet) 

    column_indexes = range(worksheet.min_column, worksheet.max_column + 1)
    for index_column, column in zip(column_indexes, worksheet.iter_cols()):

        column_width = get_column_width(column)
        dimension_holder[get_column_letter(index_column)] = ColumnDimension(worksheet, 
                                                           min=index_column, 
                                                           max=index_column,
                                                           width=column_width)
    worksheet.column_dimensions = dimension_holder

class StyleCached:
    def __init__(self,
                 font: Font,
                 border: Border,
                 fill: Fill,
                 number_format: str,
                 alignment: Alignment):
        self.font = copy(font)
        self.border = copy(border)
        self.fill = copy(fill)
        self.number_format = copy(number_format)
        self.alignment = copy(alignment)

    def __eq__(self, cell: Cell):
       return (cell.fill.__eq__(self.fill)
               and cell.font.__eq__(self.font)
               and cell.alignment.__eq__(self.alignment)
               and cell.number_format.__eq__(self.number_format)
               # 1 border all the time so not needed to check the line below
               # and cell.border.__eq__(self.border) 
        )

style_cache = []
def get_style_cached(cell: Cell):
    for cache in style_cache:
        if cache.__eq__(cell):
            return cache
    cache = StyleCached(cell.font, cell.border, cell.fill, cell.number_format, cell.alignment)
    style_cache.append(cache)
    return cache

def add_styles_to_worksheet(worksheet: Worksheet):
    logger.info(f"Adding {len(style_cache)} styles into '{worksheet.title}' worksheet...")
    for id, cache in enumerate(style_cache, 1):
        new_cell = worksheet.cell(row=id, column=1, value="style")
        new_cell.font = cache.font
        new_cell.border = cache.border
        new_cell.fill = cache.fill
        new_cell.number_format = cache.number_format
        new_cell.alignment = cache.alignment

# code inspired from https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/copier.html
def copy_sheet_to_workbook(source_sheet: Worksheet, target_workbook: Workbook) -> Worksheet:
    """
    Copy the source worksheet as a new worksheet in the target workbook
    :param source_sheet: the source sheet
    :param target_workbook: the workbook used to store the new sheet
    :return: a reference to the created sheet inside the workbook
    """
    logger.info(f"Copying sheet '{source_sheet.title}' to target workbook ({source_sheet.max_column} columns; {source_sheet.max_row} rows)...")
    target_sheet = target_workbook.create_sheet(source_sheet.title)
    for row in source_sheet:
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.data_type = cell.data_type
            if cell.has_style:
                cache = get_style_cached(cell)
                new_cell.font = cache.font
                new_cell.border = cache.border
                new_cell.fill = cache.fill
                new_cell.number_format = cache.number_format
                new_cell.alignment = cache.alignment

    return target_sheet

def rename_too_long_dataset_names(input_dataset_names: List[str]) -> Dict[str, str]:
    """
    Excel allows for only maximum 30 chars in the sheet names, so if some DS have more than 30 chars :
        - truncate the name to 28 chars
        - Add an index from 00 to 99 at the end in case of overlap
    :param input_dataset_names: the list of dataset names to remap
    :returns: a Dict[str, str] mapping the DS names with the sheet names
    """

    return_map = {}
    index_rename = -1
    renaming_length = EXCEL_MAX_LEN_SHEET_NAME - 2
    for name in input_dataset_names:
        if len(name) > EXCEL_MAX_LEN_SHEET_NAME:
            index_rename += 1
            rename = f"{name[0:renaming_length]}{index_rename:02d}"
            # Almost impossible case : a DS already has this name
            while rename in input_dataset_names:
                index_rename += 1
                rename = f"{name[0:renaming_length]}{index_rename:02d}"

            logger.info(f"Dataset '{name}' with a too long name will be stored as sheet '{rename}'")
            return_map[name] = rename
        else:
            return_map[name] = name

    return return_map

def datasets_to_xlsx(input_dataset_names, xlsx_abs_path, worksheet_provider):
    """
    Write each input dataset into one temporary excel file and merge all these excel files into the final excel file
    :param input_dataset_names: the list of dataset, using one temporary workbook per dataset
    :param xlsx_abs_path: the temporary path where to write the final excel file
    :param worksheet_provider: a lambda used to get the dataset worksheet
    """

    logger.info(f"Building output excel file '{xlsx_abs_path}'...")

    template_workbook, workbook_tmp_files = get_temporary_workbooks(input_dataset_names, worksheet_provider)

    # Save template workbook with styles and unzip it
    template_workbook_extract_dir = get_template_workbook_directory(template_workbook)

    # Move sheets into template workbook directory
    extract_and_move_temporary_worksheets_into_workbook_directory(workbook_tmp_files, template_workbook_extract_dir)

    # Build the final excel file
    logger.info("Creating the final excel file...")         
    zip_directory(template_workbook_extract_dir.name, xlsx_abs_path)
                
    print_cache()

    logger.info("Done writing output xlsx file.")

def get_temporary_workbooks(input_dataset_names, worksheet_provider):
    """
    Create a template workbook and one temporary workbook per dataset stored on disk
    :param input_dataset_names: the list of dataset, using one temporary workbook per dataset
    :param worksheet_provider: a lambda used to get the dataset worksheet
    :return a template workbook containing styles and empty workhsheets
    :return a list of temporary workbook file names (one file per dataset)
    """
    # A template workbook to store styles thanks to the cache
    template_workbook = Workbook()
    # remove the default sheet created
    template_workbook.remove(template_workbook.active)

    # List containing all temporary workbooks generated from dataset
    workbook_tmp_files = []

    renaming_map = rename_too_long_dataset_names(input_dataset_names)
    
    for name in input_dataset_names:
        dataset_worksheet = worksheet_provider(name)
        if dataset_worksheet is None:
            continue

        if name in renaming_map:
            dataset_worksheet.title = renaming_map[name]
        else:
            # should never happen
            logger.warning(f"Failed to find a name for the worksheet '{name}'")
            dataset_worksheet.title = name

        # Add an empty sheet in the template just to have the name of the dataset
        # This sheet will be replaced during the moving step
        template_workbook.create_sheet(dataset_worksheet.title)
  
        # Create a temporary workbook to save it on disk in order to avoid out of memory
        logger.info(f"Creating dataset '{name}' temporary workbook...")
        temp_workbook = Workbook()
        # Add previous styles in the default sheet (sheet1.xml) to keep indexes for the final excel file
        add_styles_to_worksheet(temp_workbook.active)
        temp_sheet = copy_sheet_to_workbook(dataset_worksheet, temp_workbook)
        logger.info(f"Styling excel sheet '{temp_sheet.title}' in temporary worksheet...")
        auto_size_column_width(temp_sheet)

        workbook_tmp_files.append(tempfile.NamedTemporaryFile())
        temp_workbook.save(workbook_tmp_files[-1].name)
        # Free memory
        del temp_sheet
        temp_workbook.close()
        del dataset_worksheet

        logger.info(f"Finished writing dataset '{name}' temporary workbook.")

    return template_workbook, workbook_tmp_files


def get_template_workbook_directory(template_workbook):
    """
    Save the template workbook with styles and unzip it into a temporary directory
    :param template_workbook: the template workbook to save and extract into a temporary directory
    :return a temporary directory
    """
    template_workbook_extract_dir = tempfile.TemporaryDirectory()
    with tempfile.NamedTemporaryFile() as template_workbook_file:
        # Add styles to template workbook before saving it
        if template_workbook.worksheets:
            add_styles_to_worksheet(template_workbook.worksheets[0])

        template_workbook.save(template_workbook_file.name)
        template_workbook.close()
        
        logger.info("Extracting template workbook...")
        with zipfile.ZipFile(template_workbook_file.name, mode="r") as zipFile:
            zipFile.extractall(path=template_workbook_extract_dir.name)

    return template_workbook_extract_dir

def extract_and_move_temporary_worksheets_into_workbook_directory(workbook_tmp_files, template_workbook_extract_dir):
    """
    Extract and move temporary worksheets into a workbook directory
    :param workbook_tmp_files: list of temporary files to extract and move
    :param template_workbook_extract_dir: the workbook directory
    """
    logger.info("Extracting and moving temporary sheets...")

    # Extract the sheet2.xml only because sheet1.xml is just for keeping style indexes
    sheet_name_to_extract_and_move = "xl/worksheets/sheet2.xml"
    for idx, file in enumerate(workbook_tmp_files, 1):
        extract_sheet_dir = tempfile.TemporaryDirectory()
        with zipfile.ZipFile(file.name, mode="r") as zipFile:
            zipFile.extract(sheet_name_to_extract_and_move, path=extract_sheet_dir.name)
        file.close() # Close file to free space disk now
        
        # Move file
        file_source = os.path.join(extract_sheet_dir.name, sheet_name_to_extract_and_move)
        file_dest = os.path.join(template_workbook_extract_dir.name, "xl/worksheets/sheet{id}.xml".format(id = idx))
        os.replace(file_source, file_dest)

def zip_directory(dir_name, output_path_file_name):
    """
    Zip a directory into an archive
    :param dir_name: the directory to zip
    :param output_path_file_name: the path file name of the archive
    """
    with zipfile.ZipFile(output_path_file_name, 'w', ZIP_DEFLATED, allowZip64=True) as archive:
        for root, dirs, files in os.walk(dir_name):
            for file in files:
                archive.write(os.path.join(root, file), 
                        os.path.relpath(os.path.join(root, file), os.path.join(dir_name, '.')))

def print_cache():
    """
    Print the counts of each style of the cache
    """
    fonts = []
    borders = []
    fills = []
    number_formats = []
    alignments = []

    def add_style_if_not_exist(style, list):
        if list:
            for s in list:
                if s.__eq__(style):
                    return
        list.append(style)

    for cache in style_cache:
        add_style_if_not_exist(cache.font, fonts)
        add_style_if_not_exist(cache.border, borders)
        add_style_if_not_exist(cache.fill, fills)
        add_style_if_not_exist(cache.alignment, alignments)
        add_style_if_not_exist(cache.number_format, number_formats)

    logger.info(f"Style counts (fonts: {len(fonts)}; borders: {len(borders)}; fills: {len(fills)}; alignments: {len(alignments)}; number_formats: {len(number_formats)})")
