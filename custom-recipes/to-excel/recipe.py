#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Custom recipe for Excel Multi Sheet Exporter
"""

import logging
import tempfile

from pathvalidate import ValidationError, validate_filename

import dataiku
from dataiku.customrecipe import get_input_names_for_role
from dataiku.customrecipe import get_output_names_for_role
from dataiku.customrecipe import get_recipe_config
from openpyxl import load_workbook, Workbook
from xlsx_writer import datasets_to_xlsx
from typing import Union

DEFAULT_DATAIKU_SHEET_NAME = "Sheet1"
READ_CHUNK_SIZE = 1024 * 1024 # 1Mbytes

def get_excel_worksheet(dataset: dataiku.Dataset, apply_conditional_formatting: bool) -> Union[Workbook, None]:
    logger.info(f"Getting Excel workbook from DSS dataset {dataset.short_name}")
    workbook = None
    with tempfile.NamedTemporaryFile(delete=True) as tmp_file:
        with dataset.raw_formatted_data(format="excel", format_params={ "applyColoring": apply_conditional_formatting }) as stream:
            # read steam with chunks to save RAM
            chunk_size = READ_CHUNK_SIZE
            while True:
                chunk = stream.read(chunk_size)
                if not chunk:
                    break
                tmp_file.write(chunk)
        tmp_file.flush() # Make sure file is written on disk
        tmp_file.seek(0) # Read back from start of file to load it in the workbook
        workbook = load_workbook(tmp_file)

    
    if workbook is not None:
        if DEFAULT_DATAIKU_SHEET_NAME in workbook:
            return workbook[DEFAULT_DATAIKU_SHEET_NAME]
        elif len(workbook.sheetnames) == 1:
            logger.warn(f"Default DSS default sheet name has changed from {DEFAULT_DATAIKU_SHEET_NAME} to {workbook.sheetnames[0]}")
            return workbook[workbook.sheetnames[0]]
    
    logger.error("Error getting Excel workbook from DSS dataset {dataset.short_name}, this dataset will not be exported")
    return None


logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='Multi-Sheet Excel Exporter | %(levelname)s - %(message)s')

input_datasets_ids = get_input_names_for_role('input_dataset')
if len(input_datasets_ids) == 0:
    logger.warning("Received no input datasets ids. input_datasets_ids={}".format(input_datasets_ids))

input_datasets_names = [name.split('.')[-1] for name in input_datasets_ids]
if len(input_datasets_names) == 0:
    logger.warning("Received no input datasets names. input_datasets_ids={}, input_datasets_names={}".format(
        input_datasets_ids, input_datasets_names))

# Retrieve the list of output folders, should contain unique element
output_folder_id = get_output_names_for_role('folder')
logger.info("Retrieved the following folder ids: {}".format(output_folder_id))
output_folder_name = output_folder_id[0]
logger.info("Received the following output folder name: {}".format(output_folder_name))
output_folder = dataiku.Folder(output_folder_name)

input_config = get_recipe_config()
workbook_name = input_config.get('output_workbook_name', None)
apply_conditional_formatting = input_config.get('export_conditional_formatting', False)

if workbook_name is None:
    logger.warning("Received input received recipe config: {}".format(input_config))
    raise ValueError('Could not read the workbook name.')

output_file_name = '{}.xlsx'.format(workbook_name)

try:
    validate_filename(output_file_name)
except ValidationError as e:
    raise ValueError(f"{e}\n")


with tempfile.NamedTemporaryFile() as tmp_file:
    tmp_file_path = tmp_file.name
    logger.info("Intend to write the output xls file to the following location: {}".format(tmp_file_path))

    datasets_to_xlsx(input_datasets_names, tmp_file_path, lambda name: get_excel_worksheet(dataiku.Dataset(name), apply_conditional_formatting))

    with open(tmp_file_path, 'rb', encoding=None) as f:
        output_folder.upload_stream(output_file_name, f)


logger.info("Ended recipe processing.")
