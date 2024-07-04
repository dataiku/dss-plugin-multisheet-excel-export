from xlsx_writer import datasets_to_xlsx, rename_too_long_dataset_names

import os
from openpyxl import Workbook
import pandas as pd
import tempfile


def build_worksheet(headers, data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in data:
        worksheet.append(row)
    return worksheet


def test_dataset_renames():
    original_list = [
        "very_very_long_name_with_too_much_character0",
        "very_very_long_name_with_too_much_character1",
        "very_very_long_name_with_too_much_character2",
        "very_very_long_name_with_too_much_character3",
        "very_very_long_name_with_too_much_character4",
        "very_very_long_name_with_too_much_character5",
        "very_very_long_name_with_too_much_character6",
        "very_very_long_name_with_too_much_character7",
        "very_very_long_name_with_too_much_character8",
        "very_very_long_name_with_too_much_character9",
        "very_very_long_name_with_too_much_character10",
        "very_very_long_name_with_too_much_character11",
        "very_very_long_name_with_too_much_character12",
        "very_very_long_name_with_too_much_character13",
        "very_very_long_name_with_too_much_character14",
        "very_very_long_name_with_too_much_character15",
        "very_very_long_name_with_too_09", # To test that too long DS doesn't get renamed to a dataset with correct len
        "very_very_long_name_with_too_14", # To test that too long DS doesn't get renamed to a dataset with correct len
        "finally_normal_dataset",
        "finally_normal_dataset2"

    ]
    test_rename_map = rename_too_long_dataset_names(original_list)
    assert test_rename_map["finally_normal_dataset"] == "finally_normal_dataset"
    assert test_rename_map["finally_normal_dataset2"] == "finally_normal_dataset2"
    assert test_rename_map["very_very_long_name_with_too_14"] == "very_very_long_name_with_too_14"
    assert test_rename_map["very_very_long_name_with_too_09"] == "very_very_long_name_with_too_09"

    assert test_rename_map["very_very_long_name_with_too_much_character0"]  == "very_very_long_name_with_too_00"
    assert test_rename_map["very_very_long_name_with_too_much_character1"]  == "very_very_long_name_with_too_01"
    assert test_rename_map["very_very_long_name_with_too_much_character2"]  == "very_very_long_name_with_too_02"
    assert test_rename_map["very_very_long_name_with_too_much_character3"]  == "very_very_long_name_with_too_03"
    assert test_rename_map["very_very_long_name_with_too_much_character4"]  == "very_very_long_name_with_too_04"
    assert test_rename_map["very_very_long_name_with_too_much_character5"]  == "very_very_long_name_with_too_05"
    assert test_rename_map["very_very_long_name_with_too_much_character6"]  == "very_very_long_name_with_too_06"
    assert test_rename_map["very_very_long_name_with_too_much_character7"]  == "very_very_long_name_with_too_07"
    assert test_rename_map["very_very_long_name_with_too_much_character8"]  == "very_very_long_name_with_too_08"
    assert test_rename_map["very_very_long_name_with_too_much_character9"]  == "very_very_long_name_with_too_10"
    assert test_rename_map["very_very_long_name_with_too_much_character10"] == "very_very_long_name_with_too_11"
    assert test_rename_map["very_very_long_name_with_too_much_character11"] == "very_very_long_name_with_too_12"
    assert test_rename_map["very_very_long_name_with_too_much_character12"] == "very_very_long_name_with_too_13"
    assert test_rename_map["very_very_long_name_with_too_much_character13"] == "very_very_long_name_with_too_15"
    assert test_rename_map["very_very_long_name_with_too_much_character14"] == "very_very_long_name_with_too_16"
    assert test_rename_map["very_very_long_name_with_too_much_character15"] == "very_very_long_name_with_too_17"




def test_datasets_to_xlsx():
    output_file_name = 'sample_test.xlsx'
    tmp_output_dir = tempfile.TemporaryDirectory(dir='.')
    output_file = os.path.join(tmp_output_dir.name, output_file_name)

    df1 = build_worksheet(['dfId', 'gender', 'birthdate'], [[1, 'M', '1953/10/5'], [2, 'L', '1053/12/6']])
    df2 = build_worksheet(['dfId', 'gender'], [[2, 'M']])
    df3 = build_worksheet(['dfId'], [[3],[4],[5],[6],[7]])
    tables = {'df1': df1, 'df2': df2, 'df3': df3}
    
    def worksheet_provider(name):
        return tables[name]

    datasets_to_xlsx(['df1', 'df2', 'df3'], output_file, worksheet_provider)

    df1_out = pd.read_excel(output_file, sheet_name="df1")
    df2_out = pd.read_excel(output_file, sheet_name="df2")
    df3_out = pd.read_excel(output_file, sheet_name="df3")
    os.remove(output_file)

    # len on a panda dataframe returns number of rows without the column names
    assert tables['df1'].max_row-1 == len(df1_out)
    assert tables['df2'].max_row-1 == len(df2_out)
    assert tables['df3'].max_row-1 == len(df3_out)

